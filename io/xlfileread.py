import sys
import xlrd
import openpyxl

def fileread(file_name, sheet_name, callbacks={}):
    if file_name.lower().endswith('.xls'):
        xlsfileread(file_name, sheet_name, callbacks)
    elif file_name.lower().endswith('.xlsx'):
        xlsxfileread(file_name, sheet_name, callbacks)
    else:
        print('not support file format!')

def call_if_defined(obj, func_name, params):
    if obj and func_name in obj and obj[func_name] and callable(obj[func_name]):
        return obj[func_name](params)
    return False


def xlsfileread(file_name, sheet_name, callbacks = {}):
    # 打开excel
    wb = xlrd.open_workbook(file_name)

    if not call_if_defined(callbacks, 'after_book_open', {'workbook': wb}):
        return

    # 设置读取编码
    # wb = xlrd.open_workbook(..., encoding_override="cp1252")

    sheets = wb.sheets()
    sheet_names = wb.sheet_names()

    for i in range(len(sheets)):
        sheet = sheets[i]
        if not call_if_defined(callbacks, 'before_sheet_open', {'sheet': sheet, 'name': sheet_names[i]}):
            continue

        for row in sheet.get_rows():
            if not call_if_defined(callbacks, 'before_row_open'
                , {'sheet': sheet, 'name': sheet_names[i], 'row': row}):
                continue
            for cell in row:
                if not call_if_defined(callbacks, 'cell_open'
                    , {'sheet': sheet, 'name': sheet_names[i], 'row': row, 'cell': cell}):
                    continue
            if not call_if_defined(callbacks, 'after_row_open'
                , {'sheet': sheet, 'name': sheet_names[i], 'row': row}):
                continue
    return

    sh = wb.sheet_by_name(sheet_name)

    # 有效数据行数
    print(sh.nrows)
    # 有效数据列数
    print(sh.ncols)
    # 输出第一行第一列的值
    print(sh.cell(0,0).value)
    # 输出第一行的所有值
    print(sh.row_values(0))

    # 将数据和标题组合成字典
    # print(
    #         dict(zip(sh.row_values(0),sh.row_values(1)))
    #     )

    # 遍历excel，打印所有数据
    for i in range(sh.nrows):
        print(sh.row_values(i))


def xlsxfileread(file_name, sheet_name, callbacks = {}):

    # 打印当前范围下的变量、方法和定义的类型列表
    # print(dir())

    # 打印a对象的属性和方法列表
    # print(dir(openpyxl))

    # 打印帮助文档
    # print(help(openpyxl))

    # 打开excel
    wb = openpyxl.load_workbook(file_name)

    if not call_if_defined(callbacks, 'after_book_open', {'workbook': wb}):
        return

    sheets = wb.worksheets
    sheet_names = wb.sheetnames
    for i in range(len(sheets)):
        sheet = sheets[i]
        if not call_if_defined(callbacks, 'before_sheet_open', {'sheet': sheet, 'name': sheet_names[i]}):
            continue

        for row in sheet.rows:
            if not call_if_defined(callbacks, 'before_row_open'
                , {'sheet': sheet, 'name': sheet_names[i], 'row': row}):
                continue
            for cell in row:
                if not call_if_defined(callbacks, 'cell_open'
                    , {'sheet': sheet, 'name': sheet_names[i], 'row': row, 'cell': cell}):
                    continue
            if not call_if_defined(callbacks, 'after_row_open'
                , {'sheet': sheet, 'name': sheet_names[i], 'row': row}):
                continue
    return

    # print(dir(wb))
    # print(wb.defined_names)
    # print(wb.defined_names.definedName)

    # print(wb.sheetnames)
    # print(wb.worksheets)

    # 索引获取sheet
    # sh = wb.sheetnames[1]
    # 索引名称获取sheet
    sh = wb[sheet_name]

    rows_generator = sh.rows
    rows = list(sh.rows)
    cols_generator = sh.columns
    cols = list(sh.columns)

    # 有效数据行数
    print(len(rows))

    # 有效数据列数
    print(len(cols))

    # 输出第一行第一列的值
    print(sh.cell(1,1).value)

    # 输出第一行的所有值
    content = ''
    for row in sh.rows:
        for cell in row:
            content += str(cell.value) + ', '
        break
    print(content)

    # 将数据和标题组合成字典
    # print(
    #         dict(zip(sh.row_values(0),sh.row_values(1)))
    #     )

    # 遍历excel，打印所有数据
    content = ''
    for row in sh.rows:
        row_list = []
        for cell in row:
            row_list.append(cell.value)
        content += str(row_list) + '\n'
    print(content)
